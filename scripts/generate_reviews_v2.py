#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
카페24 리뷰 자동 생성 스크립트 V2
실제 업로드 양식 기반
"""

import openpyxl
from datetime import datetime, timedelta
import random
import json
import sys
from openai import OpenAI
from diverse_review_templates import DIVERSE_REVIEW_TEMPLATES

# OpenAI 클라이언트 초기화 (선택사항)
try:
    import os
    api_key = os.environ.get('OPENAI_API_KEY')
    if api_key:
        client = OpenAI(api_key=api_key)
        GPT_ENABLED = True
    else:
        GPT_ENABLED = False
except Exception:
    GPT_ENABLED = False

# 다양한 말투의 리뷰 템플릿 import (48/36/35개 = 총 119개 템플릿)
REVIEW_TEMPLATES = DIVERSE_REVIEW_TEMPLATES

# 작성자 이름 생성용 성/이름 풀
LAST_NAMES = [
    "김", "이", "박", "최", "정", "강", "조", "윤", "장", "임",
    "한", "오", "서", "신", "권", "황", "안", "송", "전", "홍",
    "유", "배", "백", "허", "남", "심", "노", "하", "곽", "성",
    "차", "주", "우", "구", "신", "임", "나", "전", "민", "지"
]

FIRST_NAMES = [
    "민준", "서준", "예준", "도윤", "시우", "주원", "하준", "지호", "지후", "준서",
    "준우", "현우", "도현", "건우", "우진", "선우", "서진", "민재", "현준", "연우",
    "유준", "정우", "승우", "승현", "시윤", "승민", "지환", "승준", "유진", "지훈",
    "서연", "서윤", "지우", "서현", "민서", "하은", "하윤", "윤서", "지유", "채원",
    "지안", "수아", "소율", "예은", "다은", "예린", "수빈", "소윤", "지민", "채은",
    "서아", "예나", "채윤", "은서", "가은", "수연", "예서", "다인", "유나", "하린",
    "소연", "민지", "지영", "수진", "영희", "영숙", "정아", "미영", "현정", "은지",
    "철수", "영수", "민수", "동현", "상훈", "재훈", "태양", "성호", "진우", "상우"
]

def generate_korean_name():
    """한글 이름 3글자 자동 생성 (성 1글자 + 이름 2글자)"""
    last_name = random.choice(LAST_NAMES)
    first_name = random.choice(FIRST_NAMES)
    return f"{last_name}{first_name}"

def generate_review_with_gpt(category):
    """GPT로 리뷰 생성 (70~120자) - 매번 다른 리뷰 생성"""
    if not GPT_ENABLED:
        return None
    
    try:
        # 카테고리별 언급 포인트 (더 다양하게)
        category_points = {
            '액상': ['맛과 향', '목넘김', '가격', '재구매 의향', '질림 여부', '배송 속도', '포장 상태', '사용 기간', '향의 지속성', '목넘김 부드러움'],
            '기기': ['배터리', '디자인', '휴대성', '성능', '충전', '조작감', '내구성', '무게감', '크기', '누수 방지'],
            '일회용': ['휴대성', '맛', '사용 기간', '간편함', '가성비', '디자인', '향', '목넘김', '크기', '무게']
        }
        
        points = category_points.get(category, ['사용 경험'])
        selected_point = random.choice(points)
        
        # 말투 패턴 대폭 확장 (15가지)
        tone_variations = [
            '친근하고 캐주얼한 반말투 (예: ~네요, ~에요, ~거든요)',
            '정중하고 존댓말 (예: ~습니다, ~입니다)',
            '간결하고 직설적인 서술 (예: 좋음, 만족, 추천)',
            '경험 중심의 스토리텔링 (예: ~했는데, ~하니까)',
            '비교 중심 표현 (예: ~보다 나음, ~에 비해)',
            '감탄사 포함 표현 (예: 오, 와, 역시, 진짜)',
            '질문형 표현 섞기 (예: ~아닌가요?, ~지 않나요?)',
            '나열형 표현 (예: 첫째로, 둘째로)',
            '소박하고 꾸밈없는 표현 (예: 그냥, 딱, 쏘쏘)',
            '추천/비추천 중심 (예: 강추, 비추 아님)',
            '후회/만족 중심 (예: 잘 샀다, 후회 없다)',
            '재구매 강조 (예: 또 살 거임, 리필 예정)',
            '신뢰성 강조 (예: 믿고 샀다, 역시)',
            '가성비 중심 (예: 이 가격에, 가심비 최고)',
            '초보자 시선 (예: 처음인데, 입문자에게)'
        ]
        selected_tone = random.choice(tone_variations)
        
        # 문장 구조도 다양하게
        sentence_patterns = [
            '2문장 (짧고 간결)',
            '3문장 (중간 길이)',
            '1개 긴 문장 (나열형)'
        ]
        selected_pattern = random.choice(sentence_patterns)
        
        prompt = f"""전자담배 제품 리뷰를 작성해주세요.

카테고리: {category}
중점 언급: {selected_point}
말투 스타일: {selected_tone}
문장 구조: {selected_pattern}

**핵심**: 절대 비슷한 표현 반복 금지! 매번 완전히 다른 단어와 문장 구조 사용!

작성 조건:
- 70~120자
- 실제 사용자처럼 자연스럽게
- 긍정적 또는 중립적
- 제품명 언급 금지
- 이모지, 해시태그 금지
- 광고 같은 표현 금지
- 반복되는 상투적 표현 피하기

예시 (이런 식으로 다양하게):
1. "처음 써보는 거라 걱정했는데 생각보다 훨씬 괜찮네요. 향도 부담스럽지 않고 목넘김도 나쁘지 않아요."
2. "솔직히 기대 안 했는데 의외로 만족스럽습니다. 배터리도 오래가고 디자인도 깔끔해요."
3. "가격 대비 괜찮은 편. 맛도 나쁘지 않고 사용하기 편함. 재구매 의향 있음."

리뷰 내용만 출력하세요."""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 다양한 배경을 가진 실제 전자담배 사용자입니다. 매번 완전히 다른 말투, 표현, 단어를 사용합니다. 절대 비슷한 리뷰를 작성하지 않습니다."},
                {"role": "user", "content": prompt}
            ],
            temperature=1.5,  # 더 높은 창의성
            max_tokens=200,
            presence_penalty=1.0,  # 반복 방지 강화
            frequency_penalty=1.0  # 빈도 기반 반복 방지 강화
        )
        
        review = response.choices[0].message.content.strip()
        review = review.replace('"', '').replace("'", '').strip()
        
        # 70~120자 범위 체크
        if len(review) > 120:
            review = review[:120]
        elif len(review) < 50:  # 너무 짧으면 템플릿 사용
            return None
        
        return review
        
    except Exception as e:
        return None

def generate_title_with_gpt(category):
    """GPT로 제목 생성"""
    if not GPT_ENABLED:
        return None
    
    try:
        prompt = f"""전자담배 리뷰 제목을 작성해주세요.

카테고리: {category}

조건:
- 10~15자
- 간결하고 자연스럽게
- 긍정적 또는 중립적
- 이모지 금지

예시: "만족스러운 제품이에요", "재구매 의향 100%"

제목만 출력하세요."""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 전자담배 리뷰 작성자입니다."},
                {"role": "user", "content": prompt}
            ],
            temperature=1.0,
            max_tokens=30
        )
        
        title = response.choices[0].message.content.strip()
        title = title.replace('"', '').replace("'", '').strip()
        
        if len(title) > 20:
            title = title[:20]
        
        return title
        
    except Exception as e:
        return None

# 카테고리별 제목 템플릿
TITLE_TEMPLATES = {
    "액상": [
        "만족스러운 제품이에요",
        "재구매 의향 100%",
        "기대 이상입니다",
        "향이 정말 좋아요",
        "목넘김 부드럽네요",
        "가성비 최고",
        "품질 우수합니다",
        "강력 추천드려요",
        "정말 만족합니다",
        "다음에 또 구매할게요",
        "생각보다 훨씬 좋아요",
        "맛이 깔끔해요",
        "은은한 향이 좋네요",
        "부드럽고 좋습니다",
        "재구매 확정이에요",
        "추천합니다",
        "만족스럽네요",
        "좋은 제품이에요",
        "품질 좋아요",
        "가격도 착해요"
    ],
    "기기": [
        "디자인 깔끔해요",
        "성능 좋습니다",
        "휴대하기 편해요",
        "배터리 오래가요",
        "사용하기 편리해요",
        "가성비 좋네요",
        "입문자에게 추천",
        "만족스러운 구매",
        "품질 우수해요",
        "재구매 의향 있어요",
        "충전 빠르고 좋아요",
        "심플하고 좋습니다",
        "크기 적당해요",
        "가볍고 좋네요",
        "사용감 만족",
        "추천드려요",
        "괜찮은 제품",
        "마음에 들어요",
        "성능 만족",
        "좋은 선택이었어요"
    ],
    "일회용": [
        "간편하고 좋아요",
        "휴대성 최고",
        "맛 좋습니다",
        "편리해요",
        "출장용으로 딱",
        "여행용 추천",
        "만족스러워요",
        "가성비 좋네요",
        "간단하게 좋아요",
        "재구매 의향",
        "목넘김 부드러워요",
        "은은한 맛",
        "디자인 예뻐요",
        "사용 편해요",
        "오래가요",
        "추천합니다",
        "괜찮은 제품",
        "만족해요",
        "좋은 선택",
        "품질 좋아요"
    ]
}

def generate_title_from_content(review_content):
    """리뷰 내용을 30자 이내로 요약하여 제목 생성"""
    if not review_content:
        return "만족스러운 제품이에요"
    
    # 리뷰 내용이 30자 이하면 그대로 사용
    if len(review_content) <= 30:
        return review_content
    
    # 문장 단위로 분리
    sentences = review_content.replace('. ', '。').replace('! ', '。').replace('? ', '。').split('。')
    
    # 첫 번째 문장 사용 (30자 이내)
    first_sentence = sentences[0].strip()
    if len(first_sentence) <= 30:
        return first_sentence
    
    # 30자로 자르되, 단어 중간이 아닌 곳에서 자르기
    title = review_content[:30]
    
    # 마지막 공백 위치 찾기
    last_space = title.rfind(' ')
    if last_space > 20:  # 최소 20자는 유지
        title = title[:last_space]
    
    return title.strip()

def generate_random_datetime(days_back=3):
    """랜덤 날짜/시간 생성 (어제~3일 전, 미래 시간 방지)"""
    now = datetime.now()
    
    # 어제부터 시작 (1일 전 ~ days_back일 전)
    random_days = random.randint(1, days_back)
    random_hours = random.randint(0, 23)
    random_minutes = random.randint(0, 59)
    random_seconds = random.randint(0, 59)
    
    random_date = now - timedelta(days=random_days)
    random_date = random_date.replace(
        hour=random_hours,
        minute=random_minutes,
        second=random_seconds
    )
    
    return random_date.strftime("%Y-%m-%d %H:%M:%S")

def detect_category(product_type):
    """종류에서 카테고리 감지"""
    product_type_lower = product_type.lower()
    
    if "일회용" in product_type:
        return "일회용"
    elif "기기" in product_type or "킷" in product_type:
        return "기기"
    elif "액상" in product_type:
        return "액상"
    else:
        # 기본값은 액상
        return "액상"

def is_similar(text1, text2, threshold=0.7):
    """두 텍스트의 유사도 체크 (간단한 단어 기반)"""
    words1 = set(text1.split())
    words2 = set(text2.split())
    
    if len(words1) == 0 or len(words2) == 0:
        return False
    
    intersection = len(words1 & words2)
    union = len(words1 | words2)
    
    similarity = intersection / union if union > 0 else 0
    return similarity > threshold

def generate_reviews(products, review_count):
    """
    리뷰 생성 (중복 및 유사도 검사 강화)
    :param products: [{"product_no": "11", "type": "미션 기기", "name": "미션 전자담배 킷 - 넵튠블루"}, ...]
    :param review_count: 생성할 총 리뷰 개수
    :return: 리뷰 데이터 리스트
    """
    reviews = []
    
    # 중복 및 유사도 방지를 위한 리스트 (전체 내용 저장)
    used_reviews = []
    used_names = set()
    
    # 각 제품에 리뷰를 랜덤 분산
    for i in range(review_count):
        # 랜덤으로 제품 선택
        product = random.choice(products)
        
        # 종류에서 카테고리 감지
        category = detect_category(product.get("type", ""))
        
        # GPT로 리뷰 생성 시도 (중복 및 유사도 검사 강화)
        review_content = None
        for attempt in range(100):  # 시도 횟수 증가
            temp_review = generate_review_with_gpt(category)
            if temp_review:
                # 완전 중복 체크
                if temp_review in [r for r in used_reviews]:
                    continue
                
                # 유사도 체크 (70% 이상 유사하면 거부)
                is_too_similar = False
                for used in used_reviews:
                    if is_similar(temp_review, used, threshold=0.7):
                        is_too_similar = True
                        break
                
                if not is_too_similar:
                    review_content = temp_review
                    used_reviews.append(review_content)
                    break
        
        # GPT 실패 시 템플릿 사용 (중복 및 유사도 검사)
        if review_content is None:
            template_list = REVIEW_TEMPLATES.get(category, REVIEW_TEMPLATES["액상"])
            for attempt in range(100):  # 시도 횟수 증가
                temp_review = random.choice(template_list)
                
                # 완전 중복 체크
                if temp_review in used_reviews:
                    continue
                
                # 유사도 체크
                is_too_similar = False
                for used in used_reviews:
                    if is_similar(temp_review, used, threshold=0.7):
                        is_too_similar = True
                        break
                
                if not is_too_similar:
                    review_content = temp_review
                    used_reviews.append(review_content)
                    break
            
            # 그래도 실패하면 약간 변형하여 사용
            if review_content is None:
                base_review = random.choice(template_list)
                # 더 다양한 변형 시도
                variations = [
                    f"{base_review}",
                    f"{base_review.replace('.', '!')}",
                    f"{base_review.replace('좋아요', '좋네요')}",
                    f"{base_review.replace('만족', '흡족')}",
                    f"{base_review.replace('정말', '진짜')}",
                ]
                review_content = random.choice(variations)
                used_reviews.append(review_content)
        
        # 리뷰 내용에서 제목 생성 (30자 이내 요약)
        review_title = generate_title_from_content(review_content)
        
        # 작성자 이름 생성 (한글 이름 3글자 자동 생성, 중복 방지)
        author_name = None
        for attempt in range(50):
            temp_name = generate_korean_name()
            if temp_name not in used_names:
                author_name = temp_name
                used_names.add(author_name)
                break
        
        # 중복을 피할 수 없으면 그냥 사용
        if author_name is None:
            author_name = generate_korean_name()
        
        # 리뷰 데이터 생성
        review = {
            "product_no": product["product_no"],
            "product_code": "",  # 비워둠
            "option": "",  # 상품 옵션 비워둠
            "date": generate_random_datetime(),
            "author": author_name,
            "rating": 5,  # 항상 5점
            "title": review_title,  # 카테고리별 제목 자동 생성
            "content": review_content,
            "image1": "",
            "image2": "",
            "image3": "",
            "image4": ""
        }
        
        reviews.append(review)
    
    return reviews

def create_excel(reviews, output_path):
    """
    엑셀 파일 생성 (베이핑존/쥬스온 양식)
    리뷰에이드 업로드 형식: 1행 헤더, 2행 설명, 3행부터 데이터
    열 매핑: A=product_no, D=작성일, E=작성자, F=별점, G=제목, H=내용
    :param reviews: 리뷰 데이터 리스트
    :param output_path: 저장할 파일 경로
    """
    # 새 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 1행: 헤더 작성
    headers = [
        "product_no",      # A열
        "product_code",    # B열
        "상품 옵션",        # C열
        "작성일",           # D열
        "작성자",           # E열
        "별점",             # F열
        "제목",             # G열
        "내용",             # H열
        "이미지 URL 1",     # I열
        "이미지 URL 2",     # J열
        "이미지 URL 3",     # K열
        "이미지 URL 4"      # L열
    ]
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # 2행: 설명 작성 (리뷰에이드 양식 - example.xlsx 정확히 일치)
    descriptions = [
        "상품의 고유번호*\n리뷰에이드 어드민의 상품관리 탭에서 \n상품의 고유번호를 빠르게 확인할 수 있습니다. \n*product_no나 product_code 들 중\n하나만 입력해 주세요.",  # A열
        "상품코드*\n리뷰에이드 어드민의 상품관리 탭에서 \n상품코드를 빠르게 확인할 수 있습니다. \n*product_no나 product_code 들 중\n하나만 입력해 주세요.",  # B열
        "상품 옵션\n미입력시 공란으로 표시됩니다.",  # C열
        "리뷰 작성일*\nYYYY-MM-DD hh:mm:ss\n시간 미 입력 시 00:00:00으로 입력됩니다.",  # D열
        "리뷰 작성자 이름*",  # E열
        "리뷰 별점 \n1~5점의 만족도\n미 입력 시 별점 5점으로 입력됩니다.",  # F열
        "미 입력 시 리뷰 내용이 입력됩니다.\n제목 글자 수는 최대 250자로 제한됩니다. \n초과된 글자는 잘려서 보입니다.",  # G열
        "리뷰 내용*",  # H열
        "",  # I열
        "",  # J열
        "",  # K열
        ""   # L열
    ]
    
    for col_idx, desc in enumerate(descriptions, 1):
        ws.cell(row=2, column=col_idx, value=desc)
    
    # 3행부터: 리뷰 데이터 작성
    for row_idx, review in enumerate(reviews, 3):  # 3행부터 시작
        ws.cell(row=row_idx, column=1, value=review["product_no"])      # A열
        ws.cell(row=row_idx, column=2, value=review["product_code"])    # B열
        ws.cell(row=row_idx, column=3, value=review["option"])          # C열
        ws.cell(row=row_idx, column=4, value=review["date"])            # D열
        ws.cell(row=row_idx, column=5, value=review["author"])          # E열
        ws.cell(row=row_idx, column=6, value=review["rating"])          # F열
        ws.cell(row=row_idx, column=7, value=review["title"])           # G열
        ws.cell(row=row_idx, column=8, value=review["content"])         # H열
        ws.cell(row=row_idx, column=9, value=review["image1"])          # I열
        ws.cell(row=row_idx, column=10, value=review["image2"])         # J열
        ws.cell(row=row_idx, column=11, value=review["image3"])         # K열
        ws.cell(row=row_idx, column=12, value=review["image4"])         # L열
    
    # 파일 저장
    wb.save(output_path)
    return output_path

def main():
    """메인 함수 - CLI에서 JSON 입력 받기"""
    if len(sys.argv) < 2:
        print("Usage: python generate_reviews_v2.py '<JSON_DATA>'")
        sys.exit(1)
    
    try:
        # JSON 파라미터 파싱
        data = json.loads(sys.argv[1])
        products = data.get("products", [])
        review_count = data.get("count", 10)
        output_path = data.get("output", "cafe24_reviews.xlsx")
        
        if not products:
            print("Error: No products provided")
            sys.exit(1)
        
        # 리뷰 생성
        reviews = generate_reviews(products, review_count)
        
        # 엑셀 파일 생성
        output_file = create_excel(reviews, output_path)
        
        # 성공 메시지 (JSON 형식)
        result = {
            "success": True,
            "file": output_file,
            "review_count": len(reviews),
            "message": f"{len(reviews)}개의 리뷰가 생성되었습니다."
        }
        print(json.dumps(result, ensure_ascii=False))
        
    except Exception as e:
        # 에러 메시지 (JSON 형식)
        result = {
            "success": False,
            "error": str(e)
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

if __name__ == "__main__":
    main()
