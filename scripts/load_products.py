#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀에서 상품 리스트 불러오기
"""

import openpyxl
import json
import sys

def load_products_from_excel(excel_path):
    """
    엑셀 파일에서 상품 리스트 읽기
    :param excel_path: 엑셀 파일 경로
    :return: 상품 리스트
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        products = []
        
        # 2번째 행부터 읽기 (1번째는 헤더)
        for row_idx in range(2, ws.max_row + 1):
            product_no = ws.cell(row=row_idx, column=1).value
            product_type = ws.cell(row=row_idx, column=2).value
            product_name = ws.cell(row=row_idx, column=3).value
            
            # 유효한 데이터만 추가
            if product_no and product_type and product_name:
                products.append({
                    "product_no": str(product_no),
                    "type": str(product_type),
                    "name": str(product_name)
                })
        
        return products
        
    except Exception as e:
        raise Exception(f"엑셀 파일 읽기 실패: {e}")

def main():
    """메인 함수"""
    if len(sys.argv) < 2:
        print("Usage: python load_products.py <EXCEL_PATH>")
        sys.exit(1)
    
    try:
        excel_path = sys.argv[1]
        products = load_products_from_excel(excel_path)
        
        # JSON 형식으로 출력
        result = {
            "success": True,
            "products": products,
            "count": len(products)
        }
        print(json.dumps(result, ensure_ascii=False))
        
    except Exception as e:
        result = {
            "success": False,
            "error": str(e)
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

if __name__ == "__main__":
    main()
