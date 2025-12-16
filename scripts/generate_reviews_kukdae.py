#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
국대쥬스 전용 리뷰 생성 스크립트
엑셀 양식이 다른 국대쥬스 채널을 위한 커스텀 생성기
"""

import sys
import json
import random
import openpyxl
from datetime import datetime, timedelta
from openai import OpenAI

# OpenAI 클라이언트 초기화 (선택사항)
try:
    import os
    api_key = os.environ.get('OPENAI_API_KEY')
    if api_key:
        client = OpenAI(api_key=api_key)
        GPT_ENABLED = True
    else:
        GPT_ENABLED = False
except Exception:
    GPT_ENABLED = False

# 작성자 이름 생성용 성/이름 풀
LAST_NAMES = [
    "김", "이", "박", "최", "정", "강", "조", "윤", "장", "임",
    "한", "오", "서", "신", "권", "황", "안", "송", "전", "홍",
    "유", "배", "백", "허", "남", "심", "노", "하", "곽", "성",
    "차", "주", "우", "구", "신", "임", "나", "전", "민", "지"
]

FIRST_NAMES = [
    "민준", "서준", "예준", "도윤", "시우", "주원", "하준", "지호", "지후", "준서",
    "준우", "현우", "도현", "건우", "우진", "선우", "서진", "민재", "현준", "연우",
    "유준", "정우", "승우", "승현", "시윤", "승민", "지환", "승준", "유진", "지훈",
    "서연", "서윤", "지우", "서현", "민서", "하은", "하윤", "윤서", "지유", "채원",
    "지안", "수아", "소율", "예은", "다은", "예린", "수빈", "소윤", "지민", "채은",
    "서아", "예나", "채윤", "은서", "가은", "수연", "예서", "다인", "유나", "하린",
    "소연", "민지", "지영", "수진", "영희", "영숙", "정아", "미영", "현정", "은지",
    "철수", "영수", "민수", "동현", "상훈", "재훈", "태양", "성호", "진우", "상우"
]

def generate_korean_name():
    """한글 이름 3글자 자동 생성 (성 1글자 + 이름 2글자)"""
    last_name = random.choice(LAST_NAMES)
    first_name = random.choice(FIRST_NAMES)
    return f"{last_name}{first_name}"

# 템플릿 리뷰 (40~50자, 긍정적, 자연스러운 표현 - 매우 다양하게)
REVIEW_TEMPLATES = {
    '액상': [
        '맛 괜찮고 가격도 착한편이에요 재구매 의향 있습니다',
        '향 좋고 목넘김도 부드러워서 만족스럽게 사용중이에요',
        '생각보다 훨씬 나은것같아요 품질도 괜찮고 가성비 좋네요',
        '무난하게 쓸만한 제품이에요 가격대비 만족스럽습니다',
        '갠적으로 마음에 들어요 향도 좋고 목넘김 부드러워요',
        '가격대비 괜찮은 편이에요 맛도 좋고 질리지 않네요',
        '품질 괜찮고 가성비도 좋아요 다음에도 구매할것같아요',
        '입에 딱 맞는 맛이네요 향도 은은하고 부담없어요',
        '다쓰면 또 구매할게요 만족스럽게 잘 쓰고있습니다',
        '가성비 정말 좋은것같아요 품질도 괜찮고 만족해요',
        '재구매 고민중이에요 맛도 괜찮고 가격도 착한편이라',
        '딱히 불만은 없어요 무난하게 잘 사용하고 있습니다',
        '뭐 이정도면 만족스러워요 가격대비 괜찮은 제품이네요',
        '친구가 쓰길래 샀는데 괜찮네요 생각보다 맛 좋아요',
        '향 진하지않아서 부담없이 쓰기 좋아요 재구매 예정',
        '무난무난 쓸만한 제품이에요 가격도 착하고 품질 괜찮아요',
        '목넘김 부드럽고 맛도 좋아요 만족스럽게 사용중입니다',
        '재구매 예정이에요 품질도 괜찮고 가성비 좋은것같아요',
        '타격감 적당하고 맛도 괜찮아요 향도 은은해서 좋네요',
        '질리지 않는 맛이라 좋아요 계속 쓸것같은 제품이에요',
        '용량 대비 가격 착한편이에요 품질도 괜찮고 만족해요',
        '향이 은은해서 좋네요 목넘김도 부드럽고 가성비 좋아요',
        '처음 써보는데 나쁘지않아요 생각보다 맛있고 괜찮네요',
        '전에꺼보다 나은것같아요 향도 좋고 품질 만족스러워요',
        '담백한 맛이 좋아요 부담없이 쓰기 좋은 제품인것같아요',
    ],
    '일회용': [
        '휴대 편하고 맛도 괜찮아요 간편하게 쓰기 좋은 제품이에요',
        '간편해서 재구매 예정이에요 생각보다 오래가고 맛도 좋아요',
        '간편하게 쓰기 좋네요 휴대성도 좋고 맛도 만족스러워요',
        '생각보다 오래가요 가격대비 괜찮은 제품인것같아요',
        '가격 착한편이에요 일회용치고 품질도 괜찮고 만족해요',
        '들고다니기 편해서 좋아요 맛도 괜찮고 오래가는편이에요',
        '일회용치고 괜찮은듯해요 맛도 좋고 휴대성도 우수해요',
        '흡입감 괜찮아요 목넘김도 부드럽고 맛도 만족스러워요',
        '일주일정도 쓴것같네요 생각보다 오래가고 품질 좋아요',
        '맛도 괜찮고 가벼워서 좋아요 휴대하기 편한 제품이에요',
        '전에 쓰던거보다 나은편이에요 맛도 좋고 오래가네요',
        '휴대성 좋고 맛도 준수해요 간편하게 쓰기 딱 좋아요',
        '간편하게 쓰려고 샀는데 만족이에요 품질도 괜찮네요',
        '생각보다 괜찮아서 놀랐어요 맛도 좋고 오래가는편이에요',
        '가성비 괜찮네요 일회용치고 품질도 좋고 만족스러워요',
        '무난무난 쓸만한 제품이에요 휴대성 좋고 맛도 괜찮아요',
        '일회용으로 괜찮은 제품이에요 맛도 좋고 가격도 착해요',
        '맛 괜찮고 오래가는편이에요 가격대비 만족스러운 제품',
        '가격대비 만족스러워요 품질도 괜찮고 휴대하기 편해요',
        '버리기 아까울정도로 오래가요 맛도 좋고 품질 만족해요',
    ],
    '기기': [
        '배터리 오래가서 좋네요 충전도 빠르고 사용감 만족해요',
        '디자인 깔끔하고 성능도 괜찮아요 휴대하기도 편한 사이즈',
        '무게 가볍고 크기도 적당해요 손에 딱 맞는 그립감이에요',
        '버튼감 좋고 조작 쉬워요 입문용으로 추천할만한 제품',
        '충전 빠르고 쓰기 편해요 배터리도 오래가고 만족스러워요',
        '전에꺼보다 훨씬 나은듯해요 성능도 좋고 디자인 깔끔해요',
        '가성비 좋은 기기네요 품질도 괜찮고 사용감 만족스러워요',
        '손에 딱 맞는 크기에요 휴대성 좋고 디자인도 마음에 들어요',
        '출력 적당하고 배터리 괜찮아요 무게도 가볍고 휴대 편해요',
        '디자인 마음에들고 가벼워요 성능도 괜찮고 만족스러워요',
        '첫 기기로 쓰기 좋은듯해요 조작 간단하고 성능 괜찮아요',
        '발열 적고 성능 준수해요 배터리도 오래가고 품질 좋아요',
        '휴대성 좋고 튼튼해보여요 디자인도 깔끔하고 만족스러워요',
        '입문용으로 괜찮네요 사용법 간단하고 품질도 좋은편이에요',
        '간편하게 쓰기 좋아요 성능도 괜찮고 휴대하기 편한 사이즈',
        '가격대비 만족스러워요 품질도 괜찮고 배터리 오래가네요',
        '크기 작아서 들고다니기 편해요 성능도 좋고 만족스러워요',
        '조작법 직관적이에요 초보자도 쉽게 쓸수있는 제품인듯',
        '색상 예쁘고 실용적이에요 성능도 괜찮고 디자인 마음에 들어요',
        '배터리 용량 넉넉한편이에요 충전도 빠르고 오래가서 좋아요',
        '무게감 적당하고 그립 편해요 휴대성도 좋고 만족스러워요',
        '버튼 위치 좋고 누르기 편해요 조작감도 좋고 성능 괜찮아요',
        '충전 케이블 호환성 좋아요 배터리도 오래가고 품질 괜찮아요',
        '내구성 괜찮아보여요 튼튼하게 만들어진것같고 만족스러워요',
        '디스플레이 보기 편해요 조작도 쉽고 사용감 만족스러워요',
    ],
    '팟': [
        '누수 없어서 다행이에요 용량도 넉넉하고 맛도 좋아요',
        '교체 간편하고 맛도 좋아요 기기랑 호환성도 완벽해요',
        '용량 넉넉한 편이에요 맛도 잘 유지되고 만족스러워요',
        '맛 잘 유지되는듯해요 누수도 없고 품질 괜찮은편이에요',
        '가격대비 괜찮네요 품질도 좋고 오래 쓸수있을것같아요',
        '호환성 좋아요 기기랑 잘 맞고 교체하기도 편한 제품',
        '교체주기 적당해요 용량도 넉넉하고 맛도 오래가는편이에요',
        '누수 전혀 없어요 품질 좋고 맛도 잘 유지되는것같아요',
        '맛 오래가서 좋네요 용량 대비 가격도 착하고 만족해요',
        '기기랑 잘 맞아요 호환성 좋고 누수 걱정 없는 제품이에요',
        '품질 준수한편이에요 가격대비 괜찮고 맛도 좋은것같아요',
        '가성비 괜찮은 팟이에요 누수 없고 품질도 만족스러워요',
        '재구매 의사 있어요 맛도 좋고 용량도 넉넉한편이라',
        '전보다 맛 더 좋아요 누수도 없고 품질 향상된것같아요',
        '누수 걱정 없어요 맛도 오래가고 가격대비 만족스러워요',
        '용량 적당하고 맛도 좋아요 교체하기 편하고 품질 괜찮아요',
        '교체하기 편해요 호환성 좋고 누수 없어서 만족스러워요',
        '오래 쓸 수 있을듯해요 품질도 괜찮고 맛도 좋은편이에요',
        '무난하게 쓸만한 제품이에요 가격도 착하고 품질 괜찮아요',
        '팟치고 괜찮은 제품이에요 누수 없고 맛도 만족스러워요',
    ],
    '드립팁': [
        '입에 착 감기네요 크기도 적당하고 착용감 만족스러워요',
        '위생적이고 좋아요 세척하기 편하고 재질도 괜찮은편이에요',
        '교체하니 느낌 달라요 입술에 편하고 사용감 좋은것같아요',
        '크기 적당해요 입에 잘 맞고 착용감도 만족스러운 제품',
        '재질 좋은듯해요 내구성 괜찮아보이고 튼튼한편인것같아요',
        '입술에 편해요 자극 없고 크기도 딱 맞아서 만족스러워요',
        '세척하기 편해요 위생관리 쉽고 재질도 좋은편인것같아요',
        '내구성 괜찮아보여요 튼튼한편이고 오래 쓸수있을것같아요',
        '디자인 심플하고 좋아요 실용적이고 착용감도 괜찮은편',
        '가격 저렴해서 좋네요 품질대비 가성비 좋은 제품이에요',
        '교체주기 적당해요 내구성 괜찮고 위생적으로 관리 편해요',
        '착용감 좋아요 입에 잘 맞고 입술 자극도 없어서 좋아요',
        '입에 잘 맞아요 크기 적당하고 착용감 편한 제품인것같아요',
        '위생관리 편해요 세척 쉽고 재질도 좋아서 만족스러워요',
        '튼튼한 편이에요 내구성 괜찮고 오래 쓸수있을것같아요',
        '열전도 적당해요 입술 자극 없고 사용감도 괜찮은편이에요',
        '모양 예쁘고 실용적이에요 디자인도 좋고 착용감 만족해요',
        '가성비 좋은 드립팁이에요 품질도 괜찮고 가격도 착해요',
        '무난하게 쓸만한 제품이에요 크기 적당하고 품질 괜찮아요',
        '입술 자극 없어요 착용감 편하고 크기도 딱 맞는것같아요',
    ],
    '케이스': [
        '휴대하기 편해요 크기도 적당하고 가볍게 들고다닐수있어요',
        '보호 잘 되는듯해요 재질 튼튼하고 충격 흡수 잘되는것같아요',
        '크기 딱 맞아요 기기 보관하기 좋고 디자인도 깔끔해요',
        '재질 튼튼해보여요 내구성 괜찮고 오래 쓸수있을것같아요',
        '디자인 예쁘고 실용적이에요 색상도 마음에들고 품질 좋아요',
        '가격 착한편이에요 품질대비 가성비 좋은 케이스인것같아요',
        '가볍고 휴대성 좋아요 주머니에도 쏙 들어가서 편해요',
        '충격 보호 잘 될듯해요 재질 좋고 튼튼하게 만들어진것같아요',
        '색상 마음에 들어요 디자인도 깔끔하고 실용적인 제품이에요',
        '기기 보관하기 좋아요 크기 딱 맞고 보호 기능도 충분해요',
        '내구성 괜찮아보여요 튼튼한편이고 스크래치 방지도 잘돼요',
        '들고다니기 편해요 가벼워서 휴대성 좋고 보호도 잘되는듯',
        '가성비 좋은 케이스에요 품질도 괜찮고 디자인도 깔끔해요',
        '무난하게 쓸만한 제품이에요 보호 기능 충분하고 실용적이에요',
        '보호 기능 충분해요 재질 튼튼하고 충격 흡수 잘되는것같아요',
        '디자인 깔끔해요 색상도 예쁘고 실용적으로 잘 쓰고있어요',
        '주머니에 쏙 들어가요 크기 적당하고 휴대하기 편한 제품',
        '스크래치 방지 잘돼요 재질 좋고 기기 보호에 충분한것같아요',
        '케이스치고 괜찮네요 품질도 좋고 가격대비 만족스러워요',
        '실용적이고 좋아요 디자인 깔끔하고 보호 기능도 우수해요',
    ]
}

def generate_korean_name_with_gpt():
    """GPT로 자연스러운 3글자 한국 이름 생성"""
    if not GPT_ENABLED:
        return random.choice(REVIEWER_NAMES)
    
    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 한국 이름 생성 전문가입니다."},
                {"role": "user", "content": "자연스러운 3글자 한국 이름을 1개만 생성해주세요. 성(1글자) + 이름(2글자) 형식입니다. 김, 이, 박, 최, 정, 강, 조, 윤, 장, 임 등의 흔한 성을 사용하세요. 이름만 출력하고 다른 설명은 하지 마세요."}
            ],
            temperature=1.0,
            max_tokens=10
        )
        
        name = response.choices[0].message.content.strip()
        name = name.replace('"', '').replace("'", '').strip()
        
        if len(name) == 3:
            return name
        else:
            return random.choice(REVIEWER_NAMES)
            
    except Exception as e:
        return random.choice(REVIEWER_NAMES)

def detect_category(product_type):
    """상품 타입에서 카테고리 감지"""
    ptype_lower = product_type.lower()
    
    # 팟 감지
    if '팟' in ptype_lower or 'pod' in ptype_lower or '카트리지' in ptype_lower:
        return '팟'
    
    # 드립팁 감지
    if '드립팁' in ptype_lower or 'drip tip' in ptype_lower or '드립' in ptype_lower:
        return '드립팁'
    
    # 케이스 감지
    if '케이스' in ptype_lower or 'case' in ptype_lower:
        return '케이스'
    
    # 일회용 감지
    if '일회용' in ptype_lower or 'disposable' in ptype_lower:
        return '일회용'
    
    # 기기 감지
    if '기기' in ptype_lower or '킷' in ptype_lower or 'device' in ptype_lower or 'kit' in ptype_lower:
        return '기기'
    
    # 액상 감지
    if '액상' in ptype_lower or 'liquid' in ptype_lower or 'juice' in ptype_lower:
        return '액상'
    
    return '액상'  # 기본값

def generate_review_with_gpt(product_name, category):
    """GPT로 리뷰 생성 (50자 이내) - 매번 다른 리뷰 생성"""
    if not GPT_ENABLED:
        return None
    
    try:
        # 카테고리별 다양한 언급 포인트 (매번 랜덤 선택)
        category_points = {
            '액상': [
                '맛과 향의 강도',
                '목넘김과 타격감',
                '질림 여부와 재구매 의사',
                '가격 대비 용량',
                '향의 지속성'
            ],
            '일회용': [
                '휴대성과 간편함',
                '흡입감과 연무량',
                '사용 기간과 내구성',
                '맛의 일관성',
                '가성비'
            ],
            '기기': [
                '배터리 지속시간',
                '조작의 직관성',
                '그립감과 무게',
                '디자인과 휴대성',
                '충전 속도와 발열',
                '출력 조절 기능',
                '내구성과 견고함'
            ],
            '팟': [
                '누수 여부',
                '맛 유지력',
                '용량과 교체 주기',
                '호환성',
                '가성비'
            ],
            '드립팁': [
                '착용감과 편안함',
                '위생성',
                '재질과 내구성',
                '입술 자극 여부',
                '교체 편의성'
            ],
            '케이스': [
                '보호 기능',
                '휴대 편의성',
                '재질과 내구성',
                '디자인',
                '크기 적합성'
            ]
        }
        
        # 랜덤 포인트 선택
        points = category_points.get(category, ['제품 사용 경험'])
        selected_point = random.choice(points)
        
        # 다양한 톤과 스타일 프롬프트
        tone_variations = [
            '간결하고 직설적인 표현',
            '친근하고 캐주얼한 말투',
            '담백하고 사실적인 서술',
            '경험 중심의 구체적 표현'
        ]
        selected_tone = random.choice(tone_variations)
        
        prompt = f"""전자담배 제품에 대한 실제 구매 후기를 작성해주세요.

카테고리: {category}
이번에 중점적으로 언급할 부분: {selected_point}
말투 스타일: {selected_tone}

**중요**: 매번 완전히 다른 표현과 단어를 사용하세요. 이전 리뷰와 유사한 패턴은 피하세요.

작성 조건:
- 전체 40~50자 (최소 40자, 최대 50자)
- 2문장 정도로 구체적으로
- 실제 사용자 말투 (ㅋㅋ, ㄹㅇ, 갠적으로, 나름, 뭐, 좀 등)
- 긍정적이거나 중립적인 내용만
- 제품명 언급 금지
- 이모지, 해시태그 금지
- 광고 같은 표현 금지
- 구체적인 사용 경험 포함

예시 길이: "맛 괜찮고 가격도 착한편이에요 재구매 의향 있습니다" (29자 → 더 길게!)

리뷰 내용만 출력하세요."""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 실제 전자담배 사용자입니다. 매번 완전히 다른 표현과 단어를 사용하여 다양한 리뷰를 작성합니다. 절대 비슷한 패턴을 반복하지 마세요."},
                {"role": "user", "content": prompt}
            ],
            temperature=1.2,
            max_tokens=100,
            presence_penalty=0.8,
            frequency_penalty=0.8
        )
        
        review = response.choices[0].message.content.strip()
        review = review.replace('"', '').replace("'", '').strip()
        
        # 40~50자 범위 체크
        if len(review) > 50:
            review = review[:50]
        elif len(review) < 25:  # 너무 짧으면 템플릿 사용
            return None
        
        return review
        
    except Exception as e:
        return None

def generate_review_for_product(product_name, product_type):
    """제품에 대한 리뷰 생성 (GPT 우선, 실패 시 템플릿)"""
    category = detect_category(product_type)
    
    # GPT 시도
    gpt_review = generate_review_with_gpt(product_name, category)
    if gpt_review:
        return gpt_review
    
    # 템플릿 폴백
    templates = REVIEW_TEMPLATES.get(category, REVIEW_TEMPLATES['액상'])
    return random.choice(templates)

def generate_random_datetime_3days_ago():
    """당일~3일 전 랜덤 날짜/시간 생성 (국대쥬스 형식: 2025-04-04T07:20:55)"""
    # 0~3일 전 중 랜덤 선택
    random_days_ago = random.randint(0, 3)
    target_date = datetime.now() - timedelta(days=random_days_ago)
    
    random_hour = random.randint(0, 23)
    random_minute = random.randint(0, 59)
    random_second = random.randint(0, 59)
    
    result_datetime = target_date.replace(
        hour=random_hour,
        minute=random_minute,
        second=random_second,
        microsecond=0
    )
    
    # ISO 8601 형식으로 반환 (국대쥬스 양식)
    return result_datetime.strftime('%Y-%m-%dT%H:%M:%S')

def create_kukdae_excel(products, review_count, template_path, output_path):
    """국대쥬스 전용 엑셀 파일 생성
    
    열 매핑:
    A (1): contents - 리뷰 내용
    B (2): goodsPt - 100 (무조건)
    H (8): orderProductName - 상품명
    I (9): platformProductId - 상품번호
    L (12): writerAt - 날짜 시간 (2025-04-04T07:20:55)
    N (14): writerName - 작성자 이름
    """
    
    # 템플릿 로드
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # 기존 데이터 제거 (2행부터)
    ws.delete_rows(2, ws.max_row)
    
    current_row = 2
    
    # 중복 방지를 위한 세트
    used_reviews = set()
    used_names = set()
    
    for _ in range(review_count):
        # 랜덤 상품 선택
        product = random.choice(products)
        
        # 리뷰 생성 (중복 방지)
        review_content = None
        for attempt in range(50):  # 최대 50번 시도
            temp_review = generate_review_for_product(product['name'], product['type'])
            if temp_review not in used_reviews:
                review_content = temp_review
                used_reviews.add(review_content)
                break
        
        # 중복을 피할 수 없으면 약간 변형
        if review_content is None:
            review_content = generate_review_for_product(product['name'], product['type'])
            # 끝에 공백이나 문장부호 추가로 변형
            suffix = random.choice(['', ' ', '!', '~'])
            review_content = review_content + suffix
            used_reviews.add(review_content)
        
        # 이름 생성 (중복 방지) - 항상 자동 생성 함수 사용
        writer_name = None
        for attempt in range(50):
            temp_name = generate_korean_name()  # 자동 생성 함수 사용
            
            if temp_name not in used_names:
                writer_name = temp_name
                used_names.add(writer_name)
                break
        
        # 중복을 피할 수 없으면 그냥 사용
        if writer_name is None:
            writer_name = generate_korean_name()  # 자동 생성 함수 사용
        
        # 날짜 생성
        writer_at = generate_random_datetime_3days_ago()
        
        # 데이터 입력
        ws.cell(current_row, 1).value = review_content        # A: contents
        ws.cell(current_row, 2).value = 100                   # B: goodsPt (무조건 100)
        ws.cell(current_row, 8).value = product['name']       # H: orderProductName
        ws.cell(current_row, 9).value = str(product['product_no'])  # I: platformProductId
        ws.cell(current_row, 12).value = writer_at            # L: writerAt
        ws.cell(current_row, 14).value = writer_name          # N: writerName
        
        current_row += 1
    
    # 저장
    wb.save(output_path)
    
    return output_path

def main():
    """메인 함수"""
    try:
        # JSON 입력 받기
        if len(sys.argv) < 2:
            print(json.dumps({
                'success': False,
                'error': 'JSON 입력이 필요합니다'
            }))
            sys.exit(1)
        
        input_data = json.loads(sys.argv[1])
        
        products = input_data['products']
        count = input_data['count']
        template = input_data['template']
        output = input_data['output']
        
        # 엑셀 생성
        output_file = create_kukdae_excel(products, count, template, output)
        
        # 결과 반환
        print(json.dumps({
            'success': True,
            'message': f'{count}개의 리뷰가 생성되었습니다.',
            'file_path': output_file,
            'count': count
        }, ensure_ascii=False))
        
    except Exception as e:
        print(json.dumps({
            'success': False,
            'error': str(e)
        }, ensure_ascii=False))
        sys.exit(1)

if __name__ == '__main__':
    main()

